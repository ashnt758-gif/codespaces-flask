from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from models import db, User, Role, Permission, NFA, WorkOrder, CostContract, RevenueContract, Agreement, StatutoryDocument, Vendor, Department, Customer, Party
from utils import require_role
from sqlalchemy import func
from werkzeug.security import generate_password_hash
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import pytz

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# Admin dashboard
@admin_bp.route('/dashboard')
@login_required
@require_role('admin')
def admin_dashboard():
    """Admin dashboard with statistics"""
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    total_roles = Role.query.count()
    total_permissions = Permission.query.count()
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': total_users - active_users,
        'total_roles': total_roles,
        'total_permissions': total_permissions
    }
    
    return render_template('admin/dashboard.html', stats=stats)

# ============ User Management Routes ============

@admin_bp.route('/users', methods=['GET'])
@login_required
@require_role('admin')
def user_list():
    """List all users with pagination and search"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = User.query
    
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.first_name.ilike(f'%{search}%')) |
            (User.last_name.ilike(f'%{search}%'))
        )
    
    users = query.paginate(page=page, per_page=20)
    
    return render_template('admin/user_list.html', users=users, search=search)

@admin_bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def user_create():
    """Create a new user"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        department_id = request.form.get('department_id', '').strip()
        password = request.form.get('password', '').strip()
        role_ids = request.form.getlist('roles')
        
        # Validation
        if not username or not email or not password:
            flash('Username, email, and password are required', 'danger')
            return redirect(url_for('admin.user_create'))
        
        # Check if user exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return redirect(url_for('admin.user_create'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'danger')
            return redirect(url_for('admin.user_create'))
        
        # Create user
        user = User(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            department_id=int(department_id) if department_id else None,
            is_active=True
        )
        user.set_password(password)
        
        # Assign roles
        for role_id in role_ids:
            role = Role.query.get(role_id)
            if role:
                user.roles.append(role)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'User {username} created successfully', 'success')
        return redirect(url_for('admin.user_list'))
    
    roles = Role.query.all()
    departments = Department.query.filter_by(status='Active').all()
    return render_template('admin/user_form.html', roles=roles, departments=departments, action='Create')

@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def user_edit(user_id):
    """Edit user details"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.first_name = request.form.get('first_name', '').strip()
        user.last_name = request.form.get('last_name', '').strip()
        user.email = request.form.get('email', '').strip()
        department_id = request.form.get('department_id', '').strip()
        user.department_id = int(department_id) if department_id else None
        user.is_active = request.form.get('is_active') == 'on'
        
        # Check email uniqueness
        existing_email = User.query.filter_by(email=user.email).filter(User.id != user_id).first()
        if existing_email:
            flash('Email already in use by another user', 'danger')
            return redirect(url_for('admin.user_edit', user_id=user_id))
        
        # Update roles
        role_ids = request.form.getlist('roles')
        user.roles.clear()
        for role_id in role_ids:
            role = Role.query.get(role_id)
            if role:
                user.roles.append(role)
        
        # Update password if provided
        password = request.form.get('password', '').strip()
        if password:
            user.set_password(password)
        
        db.session.commit()
        flash(f'User {user.username} updated successfully', 'success')
        return redirect(url_for('admin.user_list'))
    
    roles = Role.query.all()
    departments = Department.query.filter_by(status='Active').all()
    user_role_ids = [role.id for role in user.roles]
    
    return render_template('admin/user_form.html', 
                         user=user, 
                         roles=roles, 
                         departments=departments,
                         user_role_ids=user_role_ids,
                         action='Edit')

@admin_bp.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@require_role('admin')
def user_toggle_active(user_id):
    """Toggle user active status"""
    user = User.query.get_or_404(user_id)
    
    # Prevent deactivating the only admin
    if user.is_active:
        admin_role = Role.query.filter_by(name='admin').first()
        if admin_role and admin_role in user.roles:
            active_admins = User.query.filter_by(is_active=True).join(User.roles).filter(Role.name == 'admin').count()
            if active_admins <= 1:
                flash('Cannot deactivate the only active admin user', 'warning')
                return redirect(url_for('admin.user_list'))
    
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'activated' if user.is_active else 'deactivated'
    flash(f'User {user.username} {status}', 'success')
    return redirect(url_for('admin.user_list'))

@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def user_delete(user_id):
    """Delete a user"""
    user = User.query.get_or_404(user_id)
    username = user.username
    
    # Prevent deleting the only admin
    admin_role = Role.query.filter_by(name='admin').first()
    if admin_role and admin_role in user.roles:
        active_admins = User.query.filter_by(is_active=True).join(User.roles).filter(Role.name == 'admin').count()
        if active_admins <= 1:
            flash('Cannot delete the only active admin user', 'warning')
            return redirect(url_for('admin.user_list'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User {username} deleted successfully', 'success')
    return redirect(url_for('admin.user_list'))

# ============ Role Management Routes ============

@admin_bp.route('/roles', methods=['GET'])
@login_required
@require_role('admin')
def role_list():
    """List all roles"""
    roles = Role.query.all()
    return render_template('admin/role_list.html', roles=roles)

@admin_bp.route('/roles/<int:role_id>', methods=['GET'])
@login_required
@require_role('admin')
def role_view(role_id):
    """View role details"""
    role = Role.query.get_or_404(role_id)
    all_permissions = Permission.query.all()
    role_permission_ids = [perm.id for perm in role.permissions]
    
    return render_template('admin/role_view.html', 
                         role=role, 
                         all_permissions=all_permissions,
                         role_permission_ids=role_permission_ids)

@admin_bp.route('/roles/<int:role_id>/edit', methods=['POST'])
@login_required
@require_role('admin')
def role_edit(role_id):
    """Edit role permissions"""
    role = Role.query.get_or_404(role_id)
    
    # Update description
    role.description = request.form.get('description', '').strip()
    
    # Update permissions
    permission_ids = request.form.getlist('permissions')
    role.permissions.clear()
    for perm_id in permission_ids:
        perm = Permission.query.get(perm_id)
        if perm:
            role.permissions.append(perm)
    
    db.session.commit()
    flash(f'Role {role.name} updated successfully', 'success')
    return redirect(url_for('admin.role_list'))

# ============ Settings Routes ============

@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def settings():
    """Admin settings page"""
    if request.method == 'POST':
        # Add any global settings here
        flash('Settings updated successfully', 'success')
        return redirect(url_for('admin.settings'))
    
    return render_template('admin/settings.html')

# ============ Reports Routes ============

@admin_bp.route('/reports')
@login_required
def reports():
    """Reports page showing all documents with filtering options"""
    from datetime import datetime, timedelta
    
    doc_type = request.args.get('doc_type', 'all', type=str)
    status = request.args.get('status', 'all', type=str)
    period = request.args.get('period', 'all', type=str)  # all, today, week, month, quarter, year, custom
    from_date_str = request.args.get('from_date', '', type=str)
    to_date_str = request.args.get('to_date', '', type=str)
    page = request.args.get('page', 1, type=int)
    
    # Calculate date range based on period
    today = datetime.utcnow()
    date_filter = None
    
    if period == 'today':
        date_filter = (today.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'week':
        start = today - timedelta(days=today.weekday())
        date_filter = (start.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'month':
        start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'quarter':
        month = today.month
        quarter_start_month = ((month - 1) // 3) * 3 + 1
        start = today.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'year':
        start = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'custom':
        # Handle custom date range
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            date_filter = (from_date, to_date)
        except (ValueError, TypeError):
            date_filter = None
    
    # Get all documents based on filters
    query_kwargs = {}
    if date_filter:
        query_kwargs['created_at_start'] = date_filter[0]
        query_kwargs['created_at_end'] = date_filter[1]
    
    nfa_records = NFA.query.all() if doc_type in ['all', 'nfa'] else []
    work_orders = WorkOrder.query.all() if doc_type in ['all', 'work_order'] else []
    cost_contracts = CostContract.query.all() if doc_type in ['all', 'cost_contract'] else []
    revenue_contracts = RevenueContract.query.all() if doc_type in ['all', 'revenue_contract'] else []
    agreements = Agreement.query.all() if doc_type in ['all', 'agreement'] else []
    statutory_docs = StatutoryDocument.query.all() if doc_type in ['all', 'statutory_document'] else []
    
    # Apply date filter
    if date_filter:
        nfa_records = [r for r in nfa_records if date_filter[0] <= r.created_at <= date_filter[1]]
        work_orders = [r for r in work_orders if date_filter[0] <= r.created_at <= date_filter[1]]
        cost_contracts = [r for r in cost_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        revenue_contracts = [r for r in revenue_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        agreements = [r for r in agreements if date_filter[0] <= r.created_at <= date_filter[1]]
        statutory_docs = [r for r in statutory_docs if date_filter[0] <= r.created_at <= date_filter[1]]
    
    # Filter by status if needed
    if status != 'all':
        nfa_records = [r for r in nfa_records if r.status == status]
        work_orders = [r for r in work_orders if r.status == status]
        cost_contracts = [r for r in cost_contracts if r.status == status]
        revenue_contracts = [r for r in revenue_contracts if r.status == status]
        agreements = [r for r in agreements if r.status == status]
        statutory_docs = [r for r in statutory_docs if r.status == status]
    
    # Combine all records with type information
    all_records = []
    for record in nfa_records:
        all_records.append({
            'type': 'NFA',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': record.amount
        })
    
    for record in work_orders:
        all_records.append({
            'type': 'Work Order',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': record.estimated_cost
        })
    
    for record in cost_contracts:
        all_records.append({
            'type': 'Cost Contract',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': record.contract_value
        })
    
    for record in revenue_contracts:
        all_records.append({
            'type': 'Revenue Contract',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': record.contract_value
        })
    
    for record in agreements:
        all_records.append({
            'type': 'Agreement',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': None
        })
    
    for record in statutory_docs:
        all_records.append({
            'type': 'Statutory Document',
            'id': record.id,
            'title': record.title,
            'reference': record.reference_number,
            'status': record.status,
            'created_by': record.created_by.username if record.created_by else 'N/A',
            'created_at': record.created_at,
            'amount': None
        })
    
    # Sort by created_at descending
    all_records.sort(key=lambda x: x['created_at'], reverse=True)
    
    # Statistics
    stats = {
        'total_records': len(all_records),
        'draft': len([r for r in all_records if r['status'] == 'Draft']),
        'submitted': len([r for r in all_records if r['status'] == 'Submitted']),
        'approved': len([r for r in all_records if r['status'] == 'Approved']),
        'rejected': len([r for r in all_records if r['status'] == 'Rejected'])
    }
    
    return render_template('admin/reports.html', 
                         records=all_records, 
                         doc_type=doc_type,
                         status=status,
                         period=period,
                         from_date=from_date_str,
                         to_date=to_date_str,
                         stats=stats)


@admin_bp.route('/reports/export/excel')
@login_required
@require_role('admin')
def export_reports_excel():
    """Export all records to Excel"""
    from datetime import datetime, timedelta
    
    doc_type = request.args.get('doc_type', 'all', type=str)
    status = request.args.get('status', 'all', type=str)
    period = request.args.get('period', 'all', type=str)
    from_date_str = request.args.get('from_date', '', type=str)
    to_date_str = request.args.get('to_date', '', type=str)
    
    # Calculate date range based on period
    today = datetime.utcnow()
    date_filter = None
    
    if period == 'today':
        date_filter = (today.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'week':
        start = today - timedelta(days=today.weekday())
        date_filter = (start.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'month':
        start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'quarter':
        month = today.month
        quarter_start_month = ((month - 1) // 3) * 3 + 1
        start = today.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'year':
        start = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'custom':
        # Handle custom date range
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            date_filter = (from_date, to_date)
        except (ValueError, TypeError):
            date_filter = None
    
    # Get all documents based on filters (same logic as reports route)
    nfa_records = NFA.query.all() if doc_type in ['all', 'nfa'] else []
    work_orders = WorkOrder.query.all() if doc_type in ['all', 'work_order'] else []
    cost_contracts = CostContract.query.all() if doc_type in ['all', 'cost_contract'] else []
    revenue_contracts = RevenueContract.query.all() if doc_type in ['all', 'revenue_contract'] else []
    agreements = Agreement.query.all() if doc_type in ['all', 'agreement'] else []
    statutory_docs = StatutoryDocument.query.all() if doc_type in ['all', 'statutory_document'] else []
    
    # Apply date filter
    if date_filter:
        nfa_records = [r for r in nfa_records if date_filter[0] <= r.created_at <= date_filter[1]]
        work_orders = [r for r in work_orders if date_filter[0] <= r.created_at <= date_filter[1]]
        cost_contracts = [r for r in cost_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        revenue_contracts = [r for r in revenue_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        agreements = [r for r in agreements if date_filter[0] <= r.created_at <= date_filter[1]]
        statutory_docs = [r for r in statutory_docs if date_filter[0] <= r.created_at <= date_filter[1]]
    
    # Filter by status if needed
    if status != 'all':
        nfa_records = [r for r in nfa_records if r.status == status]
        work_orders = [r for r in work_orders if r.status == status]
        cost_contracts = [r for r in cost_contracts if r.status == status]
        revenue_contracts = [r for r in revenue_contracts if r.status == status]
        agreements = [r for r in agreements if r.status == status]
        statutory_docs = [r for r in statutory_docs if r.status == status]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Documents"
    
    # Style headers
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Add headers
    headers = ['Document Type', 'Reference', 'Title', 'Status', 'Created By', 'Created Date', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    
    row = 2
    ist = pytz.timezone('Asia/Kolkata')
    
    # Add NFA records
    for record in nfa_records:
        ws.cell(row=row, column=1).value = "NFA"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = record.amount
        row += 1
    
    # Add WorkOrder records
    for record in work_orders:
        ws.cell(row=row, column=1).value = "Work Order"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = record.estimated_cost
        row += 1
    
    # Add CostContract records
    for record in cost_contracts:
        ws.cell(row=row, column=1).value = "Cost Contract"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = record.contract_value
        row += 1
    
    # Add RevenueContract records
    for record in revenue_contracts:
        ws.cell(row=row, column=1).value = "Revenue Contract"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = record.contract_value
        row += 1
    
    # Add Agreement records
    for record in agreements:
        ws.cell(row=row, column=1).value = "Agreement"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = 'N/A'
        row += 1
    
    # Add StatutoryDocument records
    for record in statutory_docs:
        ws.cell(row=row, column=1).value = "Statutory Document"
        ws.cell(row=row, column=2).value = record.reference_number
        ws.cell(row=row, column=3).value = record.title
        ws.cell(row=row, column=4).value = record.status
        ws.cell(row=row, column=5).value = record.created_by.username if record.created_by else 'N/A'
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        ws.cell(row=row, column=6).value = created_at_ist.strftime('%Y-%m-%d %H:%M:%S') if created_at_ist else 'N/A'
        ws.cell(row=row, column=7).value = 'N/A'
        row += 1
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'KSPL_Reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@admin_bp.route('/reports/export/pdf')
@login_required
@require_role('admin')
def export_reports_pdf():
    """Export all records to PDF"""
    from datetime import datetime, timedelta
    
    doc_type = request.args.get('doc_type', 'all', type=str)
    status = request.args.get('status', 'all', type=str)
    period = request.args.get('period', 'all', type=str)
    from_date_str = request.args.get('from_date', '', type=str)
    to_date_str = request.args.get('to_date', '', type=str)
    
    # Calculate date range based on period
    today = datetime.utcnow()
    date_filter = None
    
    if period == 'today':
        date_filter = (today.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'week':
        start = today - timedelta(days=today.weekday())
        date_filter = (start.replace(hour=0, minute=0, second=0, microsecond=0), datetime.utcnow())
    elif period == 'month':
        start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'quarter':
        month = today.month
        quarter_start_month = ((month - 1) // 3) * 3 + 1
        start = today.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'year':
        start = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = (start, datetime.utcnow())
    elif period == 'custom':
        # Handle custom date range
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            date_filter = (from_date, to_date)
        except (ValueError, TypeError):
            date_filter = None
    
    # Get all documents based on filters
    nfa_records = NFA.query.all() if doc_type in ['all', 'nfa'] else []
    work_orders = WorkOrder.query.all() if doc_type in ['all', 'work_order'] else []
    cost_contracts = CostContract.query.all() if doc_type in ['all', 'cost_contract'] else []
    revenue_contracts = RevenueContract.query.all() if doc_type in ['all', 'revenue_contract'] else []
    agreements = Agreement.query.all() if doc_type in ['all', 'agreement'] else []
    statutory_docs = StatutoryDocument.query.all() if doc_type in ['all', 'statutory_document'] else []
    
    # Apply date filter
    if date_filter:
        nfa_records = [r for r in nfa_records if date_filter[0] <= r.created_at <= date_filter[1]]
        work_orders = [r for r in work_orders if date_filter[0] <= r.created_at <= date_filter[1]]
        cost_contracts = [r for r in cost_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        revenue_contracts = [r for r in revenue_contracts if date_filter[0] <= r.created_at <= date_filter[1]]
        agreements = [r for r in agreements if date_filter[0] <= r.created_at <= date_filter[1]]
        statutory_docs = [r for r in statutory_docs if date_filter[0] <= r.created_at <= date_filter[1]]
    
    # Filter by status if needed
    if status != 'all':
        nfa_records = [r for r in nfa_records if r.status == status]
        work_orders = [r for r in work_orders if r.status == status]
        cost_contracts = [r for r in cost_contracts if r.status == status]
        revenue_contracts = [r for r in revenue_contracts if r.status == status]
        agreements = [r for r in agreements if r.status == status]
        statutory_docs = [r for r in statutory_docs if r.status == status]
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph("KSPL Documents Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare data for table
    data = [['Document Type', 'Reference', 'Title', 'Status', 'Created By', 'Created Date', 'Amount']]
    
    ist = pytz.timezone('Asia/Kolkata')
    
    # Add NFA records
    for record in nfa_records:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'NFA',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            str(record.amount) if record.amount else 'N/A'
        ])
    
    # Add other records similarly
    for record in work_orders:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'Work Order',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            str(record.estimated_cost) if record.estimated_cost else 'N/A'
        ])
    
    for record in cost_contracts:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'Cost Contract',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            str(record.contract_value) if record.contract_value else 'N/A'
        ])
    
    for record in revenue_contracts:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'Revenue Contract',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            str(record.contract_value) if record.contract_value else 'N/A'
        ])
    
    for record in agreements:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'Agreement',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            'N/A'
        ])
    
    for record in statutory_docs:
        created_at_ist = record.created_at.replace(tzinfo=pytz.utc).astimezone(ist) if record.created_at else None
        data.append([
            'Statutory Document',
            record.reference_number,
            record.title[:30],
            record.status,
            record.created_by.username if record.created_by else 'N/A',
            created_at_ist.strftime('%Y-%m-%d %H:%M') if created_at_ist else 'N/A',
            'N/A'
        ])
    
    # Create table
    table = Table(data, colWidths=[1.2*inch, 1.1*inch, 1.8*inch, 1*inch, 1.1*inch, 1.3*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'KSPL_Reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

# ============ Vendor Master Routes ============

@admin_bp.route('/vendors', methods=['GET'])
@login_required
@require_role('admin')
def vendor_list():
    """List all vendors"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Vendor.query
    
    if search:
        query = query.filter(
            (Vendor.name.ilike(f'%{search}%')) |
            (Vendor.code.ilike(f'%{search}%')) |
            (Vendor.email.ilike(f'%{search}%'))
        )
    
    vendors = query.paginate(page=page, per_page=20)
    return render_template('admin/vendor_list.html', vendors=vendors, search=search)

@admin_bp.route('/vendors/create', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def vendor_create():
    """Create a new vendor"""
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        name = request.form.get('name', '').strip()
        contact_person = request.form.get('contact_person', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        city = request.form.get('city', '').strip()
        state = request.form.get('state', '').strip()
        country = request.form.get('country', '').strip()
        
        # Validation
        if not code or not name:
            flash('Vendor Code and Name are required', 'danger')
            return redirect(url_for('admin.vendor_create'))
        
        # Check if vendor code exists
        if Vendor.query.filter_by(code=code).first():
            flash('Vendor Code already exists', 'danger')
            return redirect(url_for('admin.vendor_create'))
        
        # Create vendor
        vendor = Vendor(
            code=code,
            name=name,
            contact_person=contact_person,
            email=email,
            phone=phone,
            address=address,
            city=city,
            state=state,
            country=country,
            is_active=True
        )
        
        db.session.add(vendor)
        db.session.commit()
        
        flash(f'Vendor {name} created successfully', 'success')
        return redirect(url_for('admin.vendor_list'))
    
    return render_template('admin/vendor_form.html', action='Create')

@admin_bp.route('/vendors/<int:vendor_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def vendor_edit(vendor_id):
    """Edit vendor details"""
    vendor = Vendor.query.get_or_404(vendor_id)
    
    if request.method == 'POST':
        vendor.name = request.form.get('name', '').strip()
        vendor.contact_person = request.form.get('contact_person', '').strip()
        vendor.email = request.form.get('email', '').strip()
        vendor.phone = request.form.get('phone', '').strip()
        vendor.address = request.form.get('address', '').strip()
        vendor.city = request.form.get('city', '').strip()
        vendor.state = request.form.get('state', '').strip()
        vendor.country = request.form.get('country', '').strip()
        vendor.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        flash(f'Vendor {vendor.name} updated successfully', 'success')
        return redirect(url_for('admin.vendor_list'))
    
    return render_template('admin/vendor_form.html', vendor=vendor, action='Edit')

@admin_bp.route('/vendors/<int:vendor_id>/toggle', methods=['POST'])
@login_required
@require_role('admin')
def vendor_toggle_active(vendor_id):
    """Toggle vendor active status"""
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.is_active = not vendor.is_active
    db.session.commit()
    
    status = 'activated' if vendor.is_active else 'deactivated'
    flash(f'Vendor {vendor.name} {status}', 'success')
    return redirect(url_for('admin.vendor_list'))

@admin_bp.route('/vendors/<int:vendor_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def vendor_delete(vendor_id):
    """Delete a vendor"""
    vendor = Vendor.query.get_or_404(vendor_id)
    name = vendor.name
    
    db.session.delete(vendor)
    db.session.commit()
    
    flash(f'Vendor {name} deleted successfully', 'success')
    return redirect(url_for('admin.vendor_list'))


# ============ Customer Management Routes ============

@admin_bp.route('/customers', methods=['GET'])
@login_required
@require_role('admin')
def customer_list():
    """List all customers"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Customer.query
    
    if search:
        query = query.filter(
            (Customer.name.ilike(f'%{search}%')) |
            (Customer.code.ilike(f'%{search}%')) |
            (Customer.email.ilike(f'%{search}%'))
        )
    
    customers = query.paginate(page=page, per_page=20)
    return render_template('admin/customer_list.html', customers=customers, search=search)

@admin_bp.route('/customers/create', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def customer_create():
    """Create a new customer"""
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        name = request.form.get('name', '').strip()
        contact_person = request.form.get('contact_person', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        city = request.form.get('city', '').strip()
        state = request.form.get('state', '').strip()
        country = request.form.get('country', '').strip()
        
        # Validation
        if not code or not name:
            flash('Customer Code and Name are required', 'danger')
            return redirect(url_for('admin.customer_create'))
        
        # Check if customer code exists
        if Customer.query.filter_by(code=code).first():
            flash('Customer Code already exists', 'danger')
            return redirect(url_for('admin.customer_create'))
        
        # Create customer
        customer = Customer(
            code=code,
            name=name,
            contact_person=contact_person,
            email=email,
            phone=phone,
            address=address,
            city=city,
            state=state,
            country=country,
            is_active=True
        )
        
        db.session.add(customer)
        db.session.commit()
        
        flash(f'Customer {name} created successfully', 'success')
        return redirect(url_for('admin.customer_list'))
    
    return render_template('admin/customer_form.html', action='Create')

@admin_bp.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def customer_edit(customer_id):
    """Edit customer details"""
    customer = Customer.query.get_or_404(customer_id)
    
    if request.method == 'POST':
        customer.name = request.form.get('name', '').strip()
        customer.contact_person = request.form.get('contact_person', '').strip()
        customer.email = request.form.get('email', '').strip()
        customer.phone = request.form.get('phone', '').strip()
        customer.address = request.form.get('address', '').strip()
        customer.city = request.form.get('city', '').strip()
        customer.state = request.form.get('state', '').strip()
        customer.country = request.form.get('country', '').strip()
        customer.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        flash(f'Customer {customer.name} updated successfully', 'success')
        return redirect(url_for('admin.customer_list'))
    
    return render_template('admin/customer_form.html', customer=customer, action='Edit')

@admin_bp.route('/customers/<int:customer_id>/toggle', methods=['POST'])
@login_required
@require_role('admin')
def customer_toggle_active(customer_id):
    """Toggle customer active status"""
    customer = Customer.query.get_or_404(customer_id)
    customer.is_active = not customer.is_active
    db.session.commit()
    
    status = 'activated' if customer.is_active else 'deactivated'
    flash(f'Customer {customer.name} {status}', 'success')
    return redirect(url_for('admin.customer_list'))

@admin_bp.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def customer_delete(customer_id):
    """Delete a customer"""
    customer = Customer.query.get_or_404(customer_id)
    name = customer.name
    
    db.session.delete(customer)
    db.session.commit()
    
    flash(f'Customer {name} deleted successfully', 'success')
    return redirect(url_for('admin.customer_list'))


# ============ Party Management Routes ============

@admin_bp.route('/parties', methods=['GET'])
@login_required
@require_role('admin')
def party_list():
    """List all parties"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Party.query
    
    if search:
        query = query.filter(
            (Party.name.ilike(f'%{search}%')) |
            (Party.code.ilike(f'%{search}%')) |
            (Party.email.ilike(f'%{search}%'))
        )
    
    parties = query.paginate(page=page, per_page=20)
    return render_template('admin/party_list.html', parties=parties, search=search)

@admin_bp.route('/parties/create', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def party_create():
    """Create a new party"""
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        name = request.form.get('name', '').strip()
        contact_person = request.form.get('contact_person', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        city = request.form.get('city', '').strip()
        state = request.form.get('state', '').strip()
        country = request.form.get('country', '').strip()
        
        # Validation
        if not code or not name:
            flash('Party Code and Name are required', 'danger')
            return redirect(url_for('admin.party_create'))
        
        # Check if party code exists
        if Party.query.filter_by(code=code).first():
            flash('Party Code already exists', 'danger')
            return redirect(url_for('admin.party_create'))
        
        # Create party
        party = Party(
            code=code,
            name=name,
            contact_person=contact_person,
            email=email,
            phone=phone,
            address=address,
            city=city,
            state=state,
            country=country,
            is_active=True
        )
        
        db.session.add(party)
        db.session.commit()
        
        flash(f'Party {name} created successfully', 'success')
        return redirect(url_for('admin.party_list'))
    
    return render_template('admin/party_form.html', action='Create')

@admin_bp.route('/parties/<int:party_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def party_edit(party_id):
    """Edit party details"""
    party = Party.query.get_or_404(party_id)
    
    if request.method == 'POST':
        party.name = request.form.get('name', '').strip()
        party.contact_person = request.form.get('contact_person', '').strip()
        party.email = request.form.get('email', '').strip()
        party.phone = request.form.get('phone', '').strip()
        party.address = request.form.get('address', '').strip()
        party.city = request.form.get('city', '').strip()
        party.state = request.form.get('state', '').strip()
        party.country = request.form.get('country', '').strip()
        party.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        flash(f'Party {party.name} updated successfully', 'success')
        return redirect(url_for('admin.party_list'))
    
    return render_template('admin/party_form.html', party=party, action='Edit')

@admin_bp.route('/parties/<int:party_id>/toggle', methods=['POST'])
@login_required
@require_role('admin')
def party_toggle_active(party_id):
    """Toggle party active status"""
    party = Party.query.get_or_404(party_id)
    party.is_active = not party.is_active
    db.session.commit()
    
    status = 'activated' if party.is_active else 'deactivated'
    flash(f'Party {party.name} {status}', 'success')
    return redirect(url_for('admin.party_list'))

@admin_bp.route('/parties/<int:party_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def party_delete(party_id):
    """Delete a party"""
    party = Party.query.get_or_404(party_id)
    name = party.name
    
    db.session.delete(party)
    db.session.commit()
    
    flash(f'Party {name} deleted successfully', 'success')
    return redirect(url_for('admin.party_list'))


# ============ Department Management Routes ============

@admin_bp.route('/departments', methods=['GET'])
@login_required
@require_role('admin')
def department_list():
    """List all departments with pagination and search"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Department.query
    
    # Search filter
    if search:
        query = query.filter(
            (Department.name.ilike(f'%{search}%')) |
            (Department.code.ilike(f'%{search}%'))
        )
    
    departments = query.order_by(Department.created_at.desc()).paginate(page=page, per_page=10)
    return render_template('admin/department_list.html', departments=departments, search=search)

@admin_bp.route('/departments/create', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def department_create():
    """Create a new department"""
    from forms import DepartmentForm
    
    form = DepartmentForm()
    
    if form.validate_on_submit():
        # Check if department code already exists
        existing = Department.query.filter_by(code=form.code.data).first()
        if existing:
            flash('Department code already exists', 'danger')
            return redirect(url_for('admin.department_create'))
        
        department = Department(
            name=form.name.data,
            code=form.code.data,
            description=form.description.data,
            status=form.status.data
        )
        
        db.session.add(department)
        db.session.commit()
        
        flash(f'Department {department.name} created successfully', 'success')
        return redirect(url_for('admin.department_list'))
    
    return render_template('admin/department_form.html', form=form, action='Create', department=None)

@admin_bp.route('/departments/<int:department_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def department_edit(department_id):
    """Edit a department"""
    from forms import DepartmentForm
    
    department = Department.query.get_or_404(department_id)
    form = DepartmentForm()
    
    if form.validate_on_submit():
        # Check if code is being changed and if it already exists
        if form.code.data != department.code:
            existing = Department.query.filter_by(code=form.code.data).first()
            if existing:
                flash('Department code already exists', 'danger')
                return redirect(url_for('admin.department_edit', department_id=department_id))
        
        department.name = form.name.data
        department.code = form.code.data
        department.description = form.description.data
        department.status = form.status.data
        
        db.session.commit()
        
        flash(f'Department {department.name} updated successfully', 'success')
        return redirect(url_for('admin.department_list'))
    
    elif request.method == 'GET':
        form.name.data = department.name
        form.code.data = department.code
        form.description.data = department.description
        form.status.data = department.status
    
    return render_template('admin/department_form.html', form=form, action='Edit', department=department)

@admin_bp.route('/departments/<int:department_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def department_delete(department_id):
    """Delete a department"""
    department = Department.query.get_or_404(department_id)
    name = department.name
    
    db.session.delete(department)
    db.session.commit()
    
    flash(f'Department {name} deleted successfully', 'success')
    return redirect(url_for('admin.department_list'))

